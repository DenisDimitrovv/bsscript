import openpyxl

wb_sales = openpyxl.load_workbook("sales.xlsx")
wb_items = openpyxl.load_workbook("items.xlsx")

ws_sales = wb_sales["Продажби"]
ws_items = wb_items["Артикули"]

# row 2, col 1
start_row = 2
start_col = 1

sales_end_row = 317
items_end_row = 311

items_list = []
sales_list = []

for i in range(start_row, sales_end_row):
    current_cell_value = ws_sales.cell(row=i, column=start_col).value
    sales_list.append(current_cell_value)

for i in range(start_row, items_end_row):
    current_cell_value = ws_items.cell(row=i, column=start_col).value
    items_list.append(current_cell_value)

# create a map from a list
sales_dict = {}
for i in range(0, len(sales_list)):
    current_item = sales_list[i]
    if current_item not in sales_dict:
        sales_dict[current_item] = 1
    else:
        sales_dict[current_item] += 1

for key, value in sales_dict.items():
    if value != 1:
        print(key)


