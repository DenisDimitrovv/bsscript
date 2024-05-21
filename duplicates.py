import openpyxl
import os

# user chooses the sales excel file
cwd = os.getcwd()
# get the list of excel files
excel_files = []
for file in os.listdir(cwd):
    if file.endswith(".xlsx"):
        excel_files.append(file)

# choose your sales file
print("Choose your sales file")
for i in range(0, len(excel_files)):
    print(f"({i + 1}) for {excel_files[i]}")

choice = int(input(""))
sales_filename = excel_files[choice - 1]

# load the sales workbook
wb_sales = openpyxl.load_workbook(sales_filename)

# get the first worksheet 
ws_sales = wb_sales.worksheets[0]

# row 2, col 1
start_row = 2
start_col = 1
sales_end_row = ws_sales.max_row 
sales_list = []

# create a list for all sales
for i in range(start_row, sales_end_row):
    current_cell_value = ws_sales.cell(row=i, column=start_col).value
    sales_list.append(current_cell_value)

# create a map from a list
sales_dict = {}
for i in range(0, len(sales_list)):
    current_item = sales_list[i]
    if current_item not in sales_dict:
        sales_dict[current_item] = 1
    else:
        sales_dict[current_item] += 1

# print the duplicates
no_duplicates = True 
duplicates = []
for key, value in sales_dict.items():
    if value != 1:
        no_duplicates = False 
        duplicates.append(key)

if no_duplicates:
    print("There are no duplicates in this file")
else:
    print("Your duplicates are")
    [print(file) for file in duplicates]
    
input()


